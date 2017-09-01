from bs4 import BeautifulSoup as bsoup
from openpyxl import load_workbook
import pandas as pd
import numpy as np
import seaborn as sns
import matplotlib as mpl
import matplotlib.pyplot as plt
import calendar as cd
import datetime as dt
import requests as rq
import xlsxwriter
import simplejson, urllib, requests, webbrowser
import os

def api_scrape(site, year, month):
    with rq.Session() as s:
        p = s.post('https://www.statwatch.com/login', data=payload)
        # Authorized Request
        url = 'https://www.statwatch.com/api/2.3/org/SCTTAZ/site/{}/pc/1/almanac/month-summary?year={}&month={}'.format(site, str(year), str(month))
        r = s.get(url)
    json = r.json()

    for day in range(1, cd.monthrange(year, month)[1] + 1):
        # String Format Day
        if day < 10:
            day_str = '0' + str(day)
        else:
            day_str = str(day)
        
        # Read JSON
        dataSheet = [year,
                     month,
                     day,
                     json['DayStats'][day_str]['WeatherPrecip'],
                     json['DayStats'][day_str]['WeatherDescrip'],
                     json['DayStats'][day_str]['Status'],
                     json['DayStats'][day_str]['TotalCars']]
        site_hist.append(dataSheet)

def create_att(site, site_hist):
    # Convert site_hist to DataFrame
    df = pd.DataFrame(site_hist, columns=titles)
    df['Date'] = pd.to_datetime(df.Year*10000+df.Month*100+df.Day,format='%Y%m%d')
    df['MY'] = df['Date'].dt.strftime('%b-%y')
    df['Precip'] = df['Precip'].apply(pd.to_numeric, errors='ignore')
    df['Cars'] = df['Cars'].apply(pd.to_numeric, errors='ignore')
    
    # Rain Data
    df_rain = df[['MY', 'Precip']].reset_index()
    df_rain['Metric'] = 'Rainfall'
    
    # Cars Data
    df_cars = df[['MY', 'Cars']].reset_index()
    df_cars['Metric'] = 'Car Count'
    
    # Description Data
    df_desc = df.Description.value_counts().reset_index()

    # Create Workbook if it Doesn't exist
    filename = 'Region-Weather-Data.xlsx'
    if os.path.exists(cwd + '/' + filename) == False:
        workbook = xlsxwriter.Workbook(filename)
        worksheet = workbook.add_worksheet(regions[site])
        workbook.close()
    
    # Copy old Workbook
    book = load_workbook(filename)
    writer = pd.ExcelWriter(filename, engine='openpyxl')
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

    # Save Workbook with new Data
    df.to_excel(writer, site, header=True, index=True)
    writer.save()

    # Precipitation and Car Count Graph
    sns.set_style(style='darkgrid')
    fig, ax = plt.subplots()
    sns.pointplot(x='MY', y='Precip', data=df_rain, estimator=sum, ax=ax, ci=None, hue='Metric')
    ax2 = ax.twinx()
    sns.pointplot(x='MY', y='Cars', data=df_cars, estimator=sum, ax=ax2, ci=None, hue='Metric', palette='Set1')
    plt.gcf().suptitle('{} - Rainfall and Car Count'.format(regions[site]))
    plt.savefig('{} - Rainfall and Car Count'.format(regions[site]))
    plt.clf()

    # Description Countplot
    fig, ax = plt.subplots()
    sns.countplot(x='Description', data=df)
    plt.gcf().suptitle('{} - Statuses'.format(regions[site]))
    plt.savefig('{} - Statuses'.format(regions[site]))
    plt.clf()

cwd = os.getcwd()
now = dt.datetime.now()
payload = {'username': input('StatWatch Username: '), 'password': input('StatWatch Password: '), 'chain': input('StatWatch Chain: ')}
regions = {'01C': 'Fresno CA', '03C': 'La Quinta CA', '05N': 'Reno NV', '02A': 'Phoenix AZ'}

titles = ['Year',
        'Month',
        'Day',
        'Precip',
        'Description',
        'Status',
        'Cars']

site_list = ['01C', '03C', '05N', '02A']

for site in site_list:
    site_hist = []
    """
    for year in range(now.year - 1, now.year):
        print(site + ': ' + str(year))
        for month in range(now.month,13):
            api_scrape(site, year, month)
    """
    
    year = now.year
    print(site + ': ' + str(year))
    for month in range(1, now.month):
        api_scrape(site, year, month)
    create_att(site, site_hist)
